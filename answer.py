import pysolr
import openpyxl
import requests
import os

# 数据集名称
collection_name = 'tiku'
tixing_str = ["单选题", "多选题", "判断题"]
tiku_num = 270

# def init_solr():
#     print('正在建立与solr服务器的连接...')
#     solr = pysolr.Solr(f'http://localhost:8983/solr/{collection_name}', always_commit=True)
    # # 检查数据集是否存在
    # response = requests.get(f'{solr_url}/admin/collections?action=LIST')
    # if response.status_code == 200:
    #     collections = response.json()['collections']
    #     if collection_name in collections:
    #         # 创建solr客户端连接
    #         print('正在建立与solr服务器的连接...')
    #         solr = pysolr.Solr(f'http://localhost:8983/solr/{collection_name}', always_commit=True)
    #     else:
    #      # 添加新数据集
    #         response = requests.get(f'{solr_url}/admin/collections?action=CREATE&name={collection_name}&numShards=1&replicationFactor=1')
    #         if response.status_code != 200:
    #             print(f'创建数据集 {collection_name} 失败：{response.json()}')
    #         else:
    #             print(f'创建数据集 {collection_name} 成功！')
    #         # 将题库中的题目添加到solr中
    #         wb = openpyxl.load_workbook('tiku.xlsx')
    #         ws = wb.active
    #         for row in ws.iter_rows(min_row=2, values_only=True):
    #             r0, r1, r2, r3, r4, r5, r6, r7, r8, r9, r10 = map(str, row)
    #             solr.add([
    #             {"QNumber": r0, "kecheng": r1, "shiyan": r2, "changjing": r3, "AnswerType": r4, "Question": r5, "Answer": r6, "OptionA": r7, "OptionB": r8, "OptionC": r9, "OptionD": r10}
    #             ], fieldUpdates={'Question': 'set'})
    #         print('将题库数据存放数据集中成功！')
    #         print('正在建立与solr服务器的连接...')
    #         solr = pysolr.Solr(f'http://localhost:8983/solr/{collection_name}', always_commit=True)
        # if collection_name in collections:
        #     # 如果数据集存在，删除数据集
        #     response = requests.get(f'{solr_url}/admin/collections?action=DELETE&name={collection_name}')
        #     if response.status_code != 200:
        #         print(f'删除数据集 {collection_name} 失败：{response.json()}')
        #     else:
        #         print(f'删除数据集 {collection_name} 成功')

def init_solr(file_name):
    print('\n\n>>>正在建立与solr服务器的连接...')
    solr = pysolr.Solr(f'http://localhost:8983/solr/{collection_name}', always_commit=True)
    if file_name == '':
        file_name = "tiku.xlsx"
    if os.path.exists(os.path.join(os.path.dirname(__file__), file_name)):
        print(">>>监测目录中有待上传的题库，正在上传...")
        # 将题库中的题目添加到solr中
        wb = openpyxl.load_workbook(os.path.join(os.path.dirname(__file__), file_name))
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            r0, r1, r2, r3, r4, r5, r6, r7, r8, r9, r10 = map(str, row)
            # solr.add([
            # {"QNumber": r0, "CourseName": r1, "CourseExperName": r2, "UsageScenario": r3, "AnswerType": r4, "Question": r5, "Answer": r6, "OptionA": r7, "OptionB": r8, "OptionC": r9, "OptionD": r10}
            # ])
            solr.add([
            {"QNumber": r0, "CourseName": r1, "CourseExperName": r2, "UsageScenario": r3, "AnswerType": r4, "Question": r5, "Answer": r6, "OptionA": r7, "OptionB": r8, "OptionC": r9, "OptionD": r10}
            ], fieldUpdates={'Question': 'set', "AnswerType": 'set', "Answer": 'set', "OptionA": 'set', "OptionB": 'set', "OptionC": 'set', "OptionD": 'set'})
        os.remove(os.path.join(os.path.dirname(__file__), file_name), file_name)
        print(f">>>'{file_name}'成功存放到数据库中")
    return solr

def search(keyword, solr):
    # 定义查询语句，进行题目模糊搜索
    query = f'(Question:"*{keyword}*" OR OptionA:"*{keyword}*" OR OptionB:"*{keyword}*" OR OptionC:"*{keyword}*" OR OptionD:"*{keyword}*")'
    # 查询solr
    response = solr.search(query, fl=['QNumber', 'AnswerType', 'Question', 'Answer', 'OptionA', 'OptionB', 'OptionC', 'OptionD'], rows=10)
    # 搜索并返回结果
    results = []
    for hit in response:
        if len(hit) > 6:
            result = {
                'QNumber': hit['QNumber'],
                'AnswerType': hit['AnswerType'],
                'Question': hit['Question'],
                'Answer': hit['Answer'],
                'OptionA': hit['OptionA'],
                'OptionB': hit['OptionB'],
                'OptionC': hit['OptionC'],
                'OptionD': hit['OptionD'],
            }
        else:
            result = {
            'QNumber': hit['QNumber'],
            'AnswerType': hit['AnswerType'],
            'Question': hit['Question'],
            'Answer': hit['Answer'],
            'OptionA': hit['OptionA'],
            'OptionB': hit['OptionB'],
            }
        if result not in results:
            results.append(result)
    return results

if __name__ == "__main__":
    print(f"""
| ------------------------- 必过通源 ^v^ 搜题脚本v1.0 --------------------------------- |
|              欢迎使用通源题库搜题脚本，本脚本为个人学习项目                           |
|              题源集成所有19级可用的错题集，目前已收录{tiku_num}题                            |
|              如果没有搜到可用的错题属正常现象，如果你手上有大量错题欢迎联系作者       |\n
\033[1;31m    若有任何违规违法行为，作者不承担任何法律义务,如有违规，请联系作者QQ1781434602\n\033[0m
| ------------------------------------------------------------------------------------- |""")
    file_name = input('\n>>>请输入上传的题库名(如tiku,xlsx，直接回车不做任何操作)：')
    solr = init_solr(file_name)
    while True:
        keyword = input('\n>>>请输入中文关键词[支持题目/选项模糊搜索]：')
        results = search(keyword, solr)
        print('搜索结果：\n')
        if len(results) == 0:
            print('未找到相关题目，请尝试其他关键词！\n')
            print("\033[1;31;40m【注】题库中的所有错题均来源于漓江学堂后台，题库正在努力扩充中...\n找不到题目属于正常现象，如果你手上有题目，欢迎联系作者qq1781434602\033[0m")
        else:
            for result in results:
                print(f"\033[0;33;40m[{result['QNumber'][0]}][{tixing_str[int(result['AnswerType'][0])]}]\033[0m题目：", result['Question'][0])
                print('A: ', result['OptionA'])
                print('B: ',result['OptionB'])
                if result.get('OptionC') is not None:
                    print('C: ', result['OptionC'])
                if result.get('OptionD') is not None:
                    print('D: ', result['OptionD'])
                print(f"答案：\033[1;33;42m【{result['Answer'][0]}】\033[0m")
                print("")
                